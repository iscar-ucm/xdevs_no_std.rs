import subprocess #ejecuta comandos externos desde python (para ejecutar: cargo run --example main_li --release)
import re #modifica texto en Rust
import csv #para almacenar en un archivo .csv los datos
#import shutil
import time #para medir tiempo
from pathlib import Path
from datetime import datetime
from itertools import product
import pandas as pd #para usar excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#Rutas
BASE_DIR = Path(r"C:/Users/Cristina/Documents/Cris/Teleco_UPM/Cuarto/TFG/xdevs_no_std.rs")
EXAMPLES_DIR = BASE_DIR / "xdevs_devstone" / "examples"
RESULTS = BASE_DIR / "Resultados" 

#Modelo a ejecutar
MODELS = ["li", "hi", "ho"] 

#Parámetros de WIDTH y DEPTH
WIDTH_DEPTH = range(1, 101)   

#TIMEOUT
TIMEOUT = None

#Número de repeticiones para cada WIDTH=DEPTH:
REPETITIONS = 1

#Captura de los tiempos
RE_MODEL = re.compile(r"Model creation time:\s*([\d.]+(?:µs|ms|s|ns))")
RE_SIM_CREATION = re.compile(r"Simulator creation time:\s*([\d.]+(?:µs|ms|s|ns))")
RE_SIM = re.compile(r"Simulation time:\s*([\d.]+(?:µs|ms|s|ns))")

#Conversión de tiempos a microsegundos
def to_microseconds(duration_str: str) -> float:
    duration_str = duration_str.strip()
    if duration_str.endswith("ns"):
        return float(duration_str[:-2]) / 1_000
    elif duration_str.endswith("µs"):
        return float(duration_str[:-2])
    elif duration_str.endswith("ms"):
        return float(duration_str[:-2]) * 1_000
    elif duration_str.endswith("s"):
        return float(duration_str[:-1]) * 1_000_000
    return float("nan")

def update_rust_file(model: str, size: int) -> Path:
    ruta = EXAMPLES_DIR / f"main_{model}.rs"
    text = ruta.read_text(encoding="utf-8")
 
    # Reemplaza las constantes WIDTH y DEPTH
    text = re.sub(r"(const WIDTH\s*:\s*usize\s*=\s*)\d+", rf"\g<1>{size}", text)
    text = re.sub(r"(const DEPTH\s*:\s*usize\s*=\s*)\d+", rf"\g<1>{size}", text)
 
    # Reemplaza la llamada al macro generate_model!(N, M) con los nuevos valores
    text = re.sub(
        rf"(xdevs_devstone_macros::generate_{model}!\()\d+,\s*\d+(\))",
        rf"\g<1>{size}, {size}\2",
        text,
    )
  
    archivo_final = ruta  # Sobrescribimos el archivo original (Cargo lo recompila)
    archivo_final.write_text(text, encoding="utf-8")
    return archivo_final

#Ejecutamos los modelos y capturamos los tiempos 
def run_model(model: str) -> tuple[float | None, float | None, float | None]:
    cmd = ["cargo", "run", "--release", "--example", f"main_{model}"]
    try:
        result = subprocess.run(
            cmd,
            cwd=EXAMPLES_DIR,
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=TIMEOUT,
        )
        output = result.stdout + result.stderr
 
        li_model = RE_MODEL.search(output)
        li_sim_creation = RE_SIM_CREATION.search(output)
        li_sim = RE_SIM.search(output)
 
        t_li_model = to_microseconds(li_model.group(1))  if li_model  else None
        t_li_sim_creation = to_microseconds(li_sim_creation.group(1)) if li_sim_creation else None
        t_li_sim = to_microseconds(li_sim.group(1))    if li_sim    else None
 
        return t_li_model, t_li_sim_creation, t_li_sim
 
    except subprocess.TimeoutExpired:
        print(f"  ⚠  Timeout ({TIMEOUT}s) para {model} size={size}")
        return None, None, None
    except Exception as e:
        print(f"  ⚠  Error ejecutando {model}: {e}")
        return None, None, None


#EXCEL
HEADER_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
SUBHEAD_FILL  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
ALT_FILL      = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
 
 
def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill       = fill
    cell.font       = font
    cell.alignment  = CENTER
    cell.border     = THIN_BORDER
 
 
def style_cell(cell, alt=False):
    cell.font      = NORMAL_FONT
    cell.alignment = CENTER
    cell.border    = THIN_BORDER
    if alt:
        cell.fill = ALT_FILL
 
 
def build_sheet(ws, model: str):
    """Escribe las cabeceras en la hoja del modelo."""
    ws.title = model.upper()
 
    # Fila 1: título del modelo
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"Benchmark DEVStone — Modelo {model.upper()}"
    style_header(title_cell)
    ws.row_dimensions[1].height = 22
 
    # Fila 2: cabeceras de columna
    #  WIDTH | DEPTH | Rep | Model_µs | SimCr_µs | Sim_µs   (x REPETITIONS)
    #  Columnas: A=WIDTH, B=DEPTH, luego grupos de 3 por repetición
    ws["A2"] = "WIDTH"
    ws["B2"] = "DEPTH"
    style_header(ws["A2"], SUBHEAD_FILL, SUBHEAD_FONT)
    style_header(ws["B2"], SUBHEAD_FILL, SUBHEAD_FONT)
 
    col = 3  # Empieza en columna C
    for rep in range(1, REPETITIONS + 1):
        for label in [f"Model creation (µs) R{rep}",
                      f"Sim. creation (µs) R{rep}",
                      f"Simulation (µs) R{rep}"]:
            cell = ws.cell(row=2, column=col, value=label)
            style_header(cell, SUBHEAD_FILL, SUBHEAD_FONT)
            ws.column_dimensions[get_column_letter(col)].width = 24
            col += 1
 
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.row_dimensions[2].height = 18
 
    # Fila de inicio de datos
    ws.freeze_panes = "A3"
    return 3  # próxima fila libre
 
 
def write_row(ws, row: int, size: int, timings: list[tuple]):
    """
    Escribe una fila con WIDTH, DEPTH y los tiempos de cada repetición.
    timings: lista de (t_model, t_simcr, t_sim) por repetición.
    """
    alt = (row % 2 == 0)
    ws.cell(row=row, column=1, value=size)
    ws.cell(row=row, column=2, value=size)
    style_cell(ws.cell(row=row, column=1), alt)
    style_cell(ws.cell(row=row, column=2), alt)
 
    col = 3
    for t_model, t_simcr, t_sim in timings:
        for val in [t_model, t_simcr, t_sim]:
            c = ws.cell(row=row, column=col, value=val)
            style_cell(c, alt)
            if val is not None:
                c.number_format = '#,##0.00'
            col += 1

def main():
    RESULTS.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # Elimina la hoja vacía por defecto
 
    sheets      = {}
    next_rows   = {}
 
    for model in MODELS:
        ws = wb.create_sheet()
        next_rows[model] = build_sheet(ws, model)
        sheets[model]    = ws
 
    total = len(WIDTH_DEPTH) * len(MODELS)
    done  = 0
 
    for size in WIDTH_DEPTH:
        for model in MODELS:
            done += 1
            print(f"[{done}/{total}]  model={model.upper()}  WIDTH=DEPTH={size}")
 
            # Modifica el archivo Rust
            update_rust_file(model, size)
 
            # Ejecuta REPETITIONS veces
            timings = []
            for rep in range(1, REPETITIONS + 1):
                print(f"    repetición {rep}/{REPETITIONS} …", end=" ", flush=True)
                t_model, t_simcr, t_sim = run_model(model)
                timings.append((t_model, t_simcr, t_sim))
                status = "OK" if t_sim is not None else "FALLO"
                print(status)
 
            # Escribe la fila en Excel
            write_row(sheets[model], next_rows[model], size, timings)
            next_rows[model] += 1
 
    # Guarda el libro
    output_path = RESULTS / "benchmark_devstone.xlsx"
    wb.save(output_path)
    print(f"\n✅  Resultados guardados en: {output_path}")
 
 
if __name__ == "__main__":
    main()